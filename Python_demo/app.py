import openpyxl as xl

# wb = xl.load_workbook('111.xlsx')
# sheet = wb['Sheet1']
# # cell = sheet['a1']
# # cell = sheet.cell(1,1)
# # print(cell.value)
# print(f'rows: {sheet.max_row}')

# for row in range(2,sheet.max_row + 1):
#     cell = (sheet.cell(row,3))
#     corrected_price = cell.value * 0.9              
    # print(f'corrected price is {corrected_price}')
    # corrected_price_cell = sheet.cell(row, 4)
    # corrected_price_cell.value = corrected_price

# wb.save('222.xlsx')

# wb = xl.workbook()
# ws1 = wb.create_sheet("Mysheet") 
# ws.title = "Sherry"
# print(wb.sheetnames)